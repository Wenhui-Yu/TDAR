## Wenhui Yu 2020.06.02
## author @Wenhui Yu, yuwh16@mails.tsinghua.edu.cn

import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from xlrd import open_workbook
from xlutils.copy import copy

def print_params(para_name, para):
    for i in range(len(para)):
        print(para_name[i]+':  ',para[i])

def print_value(value):
    [inter, loss_rec_t, loss_rec_s, loss_domain_u, loss_domain_i, f1_max_t, f1_max_s, pre_u, pre_i, F1_t, NDCG_t, F1_s, NDCG_s] = value
    print('iter: %3d   loss_rec %7.2f  %7.2f   loss_domain %7.2f  %7.2f' %(inter, loss_rec_t, loss_rec_s, loss_domain_u, loss_domain_i), end='   ')
    print('\033[1;31;48m', end='')
    print('f1max: %.4f  %.4f' % (f1_max_t, f1_max_s), end='   ')
    print('\033[0m', end='')
    print('pre: %.4f  %.4f  ' % (pre_u, pre_i), end='  ')
    print('f1: %.4f  %.4f  ' % (F1_t[0], F1_s[0]))

#     print(F1_t, NDCG_t, end='  ')
#     print(F1_s, NDCG_s)

def save_params(para_name,para,path_excel):
    wb = Workbook( )
    table = wb.active
    table.title = 'Parameters'
    ldata = []
    for i in range(len(para_name)):
        parameter = [para_name[i]]
        parameter_value = para[i]
        if isinstance(parameter_value, list):
            for value in parameter_value:
                parameter.append(value)
        else:
            parameter.append(parameter_value)
        ldata.append(parameter)
    for i, p in enumerate(ldata):
        for j, q in enumerate(p):
            table.cell(row = i+1, column = j+1).value = q
    wb.save(path_excel)

def save_value(df_list,path_excel,first_sheet):
    excelWriter = pd.ExcelWriter(path_excel, engine='openpyxl')

    if first_sheet is False:
        workbook = load_workbook(path_excel)
        excelWriter.book = workbook
        exist_sheets = workbook.get_sheet_names()
        for df in df_list:
            if df[1] in exist_sheets:
                workbook.remove_sheet(workbook.get_sheet_by_name(df[1]))
            df[0].to_excel(excel_writer=excelWriter, sheet_name=df[1],index = True)
            excelWriter.save()
    else:
        for df in df_list:
            df[0].to_excel(excel_writer=excelWriter, sheet_name=df[1], index=True)
            excelWriter.save()
    excelWriter.close()

def df2str(df):
    df_str = ''
    for i in range(df.shape[0]):
        df_list = df.iloc[[i], :].values.tolist()
        df_list2 = [str(i) for i in df_list]
        str_temp = ''.join(df_list2)
        df_str = df_str +str_temp+','
    return df_str